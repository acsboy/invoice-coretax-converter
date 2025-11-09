from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime
from werkzeug.utils import secure_filename
import tempfile

app = Flask(__name__)
# Auto-generate secret key for session management
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def round_decimal(value):
    """Round decimal values to nearest whole number"""
    if value is None:
        return None
    try:
        return round(float(value))
    except (ValueError, TypeError):
        return value

def create_coretax_template():
    """
    Create Coretax template structure from scratch
    Returns a new workbook with Faktur and DetailFaktur sheets
    """
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Faktur sheet
    faktur_sheet = wb.create_sheet('Faktur', 0)
    
    # Row 1: NPWP Penjual header
    faktur_sheet['A1'] = 'NPWP Penjual'
    faktur_sheet['C1'] = '0012328415631000'
    
    # Row 2: Empty
    
    # Row 3: Headers
    faktur_headers = [
        'Baris', 'Tanggal Faktur', 'Jenis Faktur', 'Kode Transaksi',
        'Keterangan Tambahan', 'Dokumen Pendukung', 'Referensi', 'Cap Fasilitas',
        'ID TKU Penjual', 'NPWP.NIK Pembeli', 'Jenis ID Pembeli', 'Negara Pembeli',
        'Nomor Dokumen Pembeli', 'Nama Pembeli', 'Alamat Pembeli', 'Email Pembeli',
        'ID TKU Pembeli'
    ]
    
    for col_idx, header in enumerate(faktur_headers, start=1):
        cell = faktur_sheet.cell(3, col_idx)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Create DetailFaktur sheet
    detail_sheet = wb.create_sheet('DetailFaktur', 1)
    
    # Row 1: Headers
    detail_headers = [
        'Baris', 'Barang.Jasa', 'Kode Barang Jasa', 'Nama Barang.Jasa',
        'Nama Satuan Ukur', 'Harga Satuan', 'Jumlah Barang Jasa', 'Total Diskon',
        'DPP', 'DPP Nilai Lain', 'Tarif PPN', 'PPN', 'Tarif PPnBM', 'PPnBM'
    ]
    
    for col_idx, header in enumerate(detail_headers, start=1):
        cell = detail_sheet.cell(1, col_idx)
        cell.value = header
        cell.font = Font(bold=True)
    
    return wb

def convert_invoice_to_coretax(sample_file_path, output_file_path):
    """
    Convert invoice list to Coretax template format
    """
    sample_wb = openpyxl.load_workbook(sample_file_path)
    sample_sheet = sample_wb.active
    
    # Create template from embedded structure
    template_wb = create_coretax_template()
    
    # Get the sheets
    faktur_sheet = template_wb['Faktur']
    detail_faktur_sheet = template_wb['DetailFaktur']
    
    # Process Faktur sheet
    faktur_row = 4  # Start writing from row 4 in template
    
    for sample_row in range(2, sample_sheet.max_row + 1):
        # Copy columns A to Q (columns 1 to 17)
        for col in range(1, 18):
            value = sample_sheet.cell(sample_row, col).value
            faktur_sheet.cell(faktur_row, col).value = value
        faktur_row += 1
    
    # Process DetailFaktur sheet
    detail_row = 2  # Start writing from row 2 in template
    
    for sample_row in range(2, sample_sheet.max_row + 1):
        # Column A: Baris
        detail_faktur_sheet.cell(detail_row, 1).value = sample_sheet.cell(sample_row, 1).value
        
        # Column B: Barang.Jasa (from column R/18)
        detail_faktur_sheet.cell(detail_row, 2).value = sample_sheet.cell(sample_row, 18).value
        
        # Column C: Kode Barang Jasa - Default value '310000
        detail_faktur_sheet.cell(detail_row, 3).value = "'310000"
        
        # Map columns T-AD to D-N
        source_to_target = {
            20: 4,   # T -> D (Nama Barang.Jasa)
            21: 5,   # U -> E (Nama Satuan Ukur)
            22: 6,   # V -> F (Harga Satuan)
            23: 7,   # W -> G (Jumlah Barang Jasa)
            24: 8,   # X -> H (Total Diskon)
            25: 9,   # Y -> I (DPP)
            26: 10,  # Z -> J (DPP Nilai Lain)
            27: 11,  # AA -> K (Tarif PPN)
            28: 12,  # AB -> L (PPN)
            29: 13,  # AC -> M (Tarif PPnBM)
            30: 14,  # AD -> N (PPnBM)
        }
        
        for source_col, target_col in source_to_target.items():
            value = sample_sheet.cell(sample_row, source_col).value
            
            # Round decimal values for numeric columns
            if target_col in [6, 7, 8, 9, 10, 12, 14]:
                value = round_decimal(value)
            
            detail_faktur_sheet.cell(detail_row, target_col).value = value
        
        detail_row += 1
    
    # Save the output file
    template_wb.save(output_file_path)
    
    return output_file_path

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert():
    # Check if file was uploaded
    if 'invoice_file' not in request.files:
        flash('Please upload your invoice list', 'error')
        return redirect(url_for('index'))
    
    invoice_file = request.files['invoice_file']
    
    # Check if file is selected
    if invoice_file.filename == '':
        flash('Please select a file', 'error')
        return redirect(url_for('index'))
    
    # Validate file type
    if not allowed_file(invoice_file.filename):
        flash('Only .xlsx files are allowed', 'error')
        return redirect(url_for('index'))
    
    try:
        # Create temporary directory for processing
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save uploaded file
            invoice_path = os.path.join(temp_dir, secure_filename(invoice_file.filename))
            invoice_file.save(invoice_path)
            
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"Coretax_Import_{timestamp}.xlsx"
            output_path = os.path.join(temp_dir, output_filename)
            
            # Convert the file using embedded template
            convert_invoice_to_coretax(invoice_path, output_path)
            
            # Send the file to user
            return send_file(
                output_path,
                as_attachment=True,
                download_name=output_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    
    except Exception as e:
        flash(f'Error during conversion: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/health')
def health():
    return {'status': 'healthy'}, 200

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
